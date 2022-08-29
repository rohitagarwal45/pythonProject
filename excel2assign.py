import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
lv_sheet1 = inv_file["Sheet1"]

total_inventory_value = {}

for record in range( 2, lv_sheet1.max_row + 1 ):
    Supplier_name = lv_sheet1.cell(record, 4).value
    price = lv_sheet1.cell(record, 3).value
    inventory = lv_sheet1.cell(record, 2).value

    if lv_sheet1.cell(record , 2 ).value < 10 :
        print(f"product no {lv_sheet1.cell(record , 1).value} No of Inventory - {int(lv_sheet1.cell(record , 2).value)}" )

    if Supplier_name in total_inventory_value :
        total_inventory_value[Supplier_name] = total_inventory_value[Supplier_name] +  price * inventory
    else :
       total_inventory_value[Supplier_name] = price * inventory


print(total_inventory_value)