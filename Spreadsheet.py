from builtins import range

import  openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
workbook_sheet = inv_file["Sheet1"]
no_of_record = workbook_sheet.max_row

dictionary_excel = {}
print(no_of_record)

for record in range(2,no_of_record):

    if workbook_sheet.cell(record, 4).value in dictionary_excel:
        dictionary_excel[workbook_sheet.cell(record, 4).value] = dictionary_excel[
                                                                     workbook_sheet.cell(record, 4).value] + 1
    else:
        dictionary_excel[workbook_sheet.cell(record, 4).value] = 1

print(dictionary_excel)
